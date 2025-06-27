from flask import Blueprint, jsonify, request, send_file
from src.models.exchange import db, Inventory, TransferCompany, Remittance, Debt, GoldTransaction, CashBoxEntry, CompanyTransfer, CurrencyConversion, Amanah
from datetime import datetime
from fpdf import FPDF
import io
import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook, load_workbook
from src.models.shop_gold import ShopGold
from src.models.user import User

inventory_bp = Blueprint('inventory', __name__)

@inventory_bp.route('/inventory', methods=['GET'])
def get_inventory():
    try:
        inventory_items = Inventory.query.all()
        all_currencies = [item.item_name for item in inventory_items if item.item_type == 'currency']
        companies = TransferCompany.query.order_by(TransferCompany.order.asc()).all()
        companies_list = []
        for company in companies:
            balances = {}
            for currency in all_currencies:
                company_item = next((item for item in inventory_items if item.item_type == 'company' and item.item_name == f'{company.company_name}__{currency}'), None)
                balance = company_item.balance if company_item else 0
                balances[currency] = balance
            companies_list.append({
                'id': company.id,
                'name': company.company_name,
                'balances': balances
            })
        result = {
            'currencies': [],
            'gold': 0,
            'companies': companies_list
        }
        for item in inventory_items:
            if item.item_type == 'currency':
                result['currencies'].append({
                    'name': item.item_name,
                    'balance': item.balance
                })
            elif item.item_type == 'gold':
                result['gold'] = item.balance
        print('DEBUG: companies returned:', [(c['id'], c['name']) for c in companies_list])
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/inventory/update', methods=['POST'])
def update_inventory():
    try:
        data = request.json
        item_name = data.get('item_name')
        amount = data.get('amount', 0)
        
        inventory_item = Inventory.query.filter_by(item_name=item_name).first()
        if not inventory_item:
            inventory_item = Inventory(item_name=item_name, balance=0, item_type='currency')
            db.session.add(inventory_item)
        
        inventory_item.balance += amount
        db.session.commit()
        
        return jsonify({'success': True, 'new_balance': inventory_item.balance})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/companies/add', methods=['POST'])
def add_company():
    try:
        data = request.json
        company_name = data.get('company_name')
        if not company_name:
            return jsonify({'error': 'اسم الشركة مطلوب'}), 400
        # تحقق إذا كانت الشركة موجودة مسبقاً
        existing = TransferCompany.query.filter_by(company_name=company_name).first()
        if existing:
            return jsonify({'error': 'الشركة موجودة مسبقاً'}), 400
        # جلب أكبر قيمة order حالية
        max_order = db.session.query(db.func.max(TransferCompany.order)).scalar() or 0
        new_company = TransferCompany(company_name=company_name, order=max_order + 1)
        db.session.add(new_company)
        db.session.commit()
        return jsonify({'success': True, 'company': {'id': new_company.id, 'name': new_company.company_name}})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/reset-all', methods=['POST'])
def reset_all():
    try:
        # حذف جميع الحوالات
        Remittance.query.delete()
        Debt.query.delete()
        GoldTransaction.query.delete()
        CashBoxEntry.query.delete()
        CompanyTransfer.query.delete()
        CurrencyConversion.query.delete()
        # إعادة جميع أرصدة الجرد إلى صفر
        for item in Inventory.query.all():
            item.balance = 0
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

def reverse_arabic(text):
    return text[::-1]

def fix_arabic(text):
    reshaped = arabic_reshaper.reshape(text)
    return get_display(reshaped)

@inventory_bp.route('/export/inventory/pdf', methods=['GET'])
def export_inventory_pdf():
    try:
        inventory_items = Inventory.query.all()
        companies = TransferCompany.query.order_by(TransferCompany.order.asc()).all()
        all_currencies = [item.item_name for item in inventory_items if item.item_type == 'currency']
        company_balances = {}
        for company in companies:
            company_balances[company.company_name] = {}
            for currency in all_currencies:
                sent_total = db.session.query(db.func.sum(Remittance.amount)).filter(
                    Remittance.company_id == company.id,
                    Remittance.type == 'send',
                    Remittance.currency == currency
                ).scalar() or 0
                received_total = db.session.query(db.func.sum(Remittance.amount)).filter(
                    Remittance.company_id == company.id,
                    Remittance.type == 'receive',
                    Remittance.currency == currency
                ).scalar() or 0
                company_balances[company.company_name][currency] = sent_total - received_total
        pdf = FPDF()
        pdf.add_page()
        pdf.add_font('Amiri', '', 'src/static/fonts/Amiri-Regular.ttf', uni=True)
        pdf.set_font('Amiri', '', 14)
        pdf.cell(0, 10, fix_arabic('تقرير الجرد العام'), ln=1, align='C')
        pdf.set_font('Amiri', '', 12)
        pdf.cell(0, 10, fix_arabic('العملات:'), ln=1)
        for item in inventory_items:
            if item.item_type == 'currency':
                pdf.cell(0, 8, f"{item.item_name}: {item.balance}", ln=1)
        pdf.cell(0, 10, fix_arabic('الذهب:'), ln=1)
        for item in inventory_items:
            if item.item_type == 'gold':
                pdf.cell(0, 8, fix_arabic(f"الذهب: {item.balance} غرام"), ln=1)
        pdf.cell(0, 10, fix_arabic('أرصدة الشركات:'), ln=1)
        for company, balances in company_balances.items():
            pdf.cell(0, 8, fix_arabic(company), ln=1)
            for currency, balance in balances.items():
                pdf.cell(0, 8, f"   {currency}: {balance}", ln=1)
        pdf_output = io.BytesIO()
        pdf.output(pdf_output)
        pdf_output.seek(0)
        return send_file(pdf_output, as_attachment=True, download_name='inventory_report.pdf', mimetype='application/pdf')
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/companies/reorder', methods=['POST'])
def reorder_companies():
    try:
        data = request.json
        ids = data.get('ids', [])
        print('DEBUG: ids received for reorder:', ids)
        for idx, company_id in enumerate(ids):
            company = TransferCompany.query.get(company_id)
            if company:
                company.order = idx + 1
        db.session.commit()
        # طباعة الترتيب بعد الحفظ
        companies = TransferCompany.query.order_by(TransferCompany.order.asc()).all()
        print('DEBUG: order after commit:', [(c.id, c.order) for c in companies])
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/export/inventory/excel', methods=['GET'])
def export_inventory_excel():
    try:
        inventory_items = Inventory.query.all()
        companies = TransferCompany.query.order_by(TransferCompany.order.asc()).all()
        all_currencies = [item.item_name for item in inventory_items if item.item_type == 'currency']
        company_balances = {}
        for company in companies:
            company_balances[company.company_name] = {}
            for currency in all_currencies:
                sent_total = db.session.query(db.func.sum(Remittance.amount)).filter(
                    Remittance.company_id == company.id,
                    Remittance.type == 'send',
                    Remittance.currency == currency
                ).scalar() or 0
                received_total = db.session.query(db.func.sum(Remittance.amount)).filter(
                    Remittance.company_id == company.id,
                    Remittance.type == 'receive',
                    Remittance.currency == currency
                ).scalar() or 0
                company_balances[company.company_name][currency] = sent_total - received_total
        wb = Workbook()
        ws = wb.active
        ws.title = 'الجرد'
        ws.append(['العملات', 'الرصيد'])
        for item in inventory_items:
            if item.item_type == 'currency':
                ws.append([item.item_name, item.balance])
        ws.append([])
        ws.append(['الذهب', 'الرصيد'])
        for item in inventory_items:
            if item.item_type == 'gold':
                ws.append(['الذهب', item.balance])
        ws.append([])
        ws.append(['أرصدة الشركات'])
        ws.append(['الشركة'] + all_currencies)
        for company, balances in company_balances.items():
            row = [company] + [balances.get(currency, 0) for currency in all_currencies]
            ws.append(row)
        output = 'test_inventory.xlsx'
        wb.save(output)
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        filename = f'inventory_{today}.xlsx'
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/import/inventory/excel', methods=['POST'])
def import_inventory_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'لم يتم رفع أي ملف'}), 400
        file = request.files['file']
        wb = load_workbook(file)
        ws = wb.active
        mode = None
        for row in ws.iter_rows(values_only=True):
            if not any(row):
                continue
            if row[0] == 'العملات':
                mode = 'currencies'
                continue
            if row[0] == 'الذهب':
                mode = 'gold'
                continue
            if row[0] == 'أرصدة الشركات':
                mode = 'companies'
                continue
            if mode == 'currencies' and row[0] and isinstance(row[1], (int, float)):
                item = Inventory.query.filter_by(item_name=row[0], item_type='currency').first()
                if not item:
                    item = Inventory(item_name=row[0], item_type='currency', balance=0)
                    db.session.add(item)
                item.balance = float(row[1])
            elif mode == 'gold' and row[0] == 'الذهب' and isinstance(row[1], (int, float)):
                item = Inventory.query.filter_by(item_type='gold').first()
                if not item:
                    item = Inventory(item_name='gold', item_type='gold', balance=0)
                    db.session.add(item)
                item.balance = float(row[1])
            elif mode == 'companies' and row[0] and isinstance(row[1], str):
                # أول صف بعد 'أرصدة الشركات' هو رؤوس الأعمدة
                # الصفوف التالية: الشركة | عملة1 | عملة2 ...
                continue  # سنتعامل مع الشركات في حلقة منفصلة
        # تحديث أرصدة الشركات
        ws = wb.active
        companies_section = False
        headers = []
        for row in ws.iter_rows(values_only=True):
            if row[0] == 'أرصدة الشركات':
                companies_section = True
                continue
            if companies_section and row[0] == 'الشركة':
                headers = list(row)
                continue
            if companies_section and row[0] and headers and len(row) >= 2:
                company_name = row[0]
                for idx, currency in enumerate(headers[1:], 1):
                    if currency and isinstance(row[idx], (int, float)):
                        # تحديث أو إنشاء رصيد الشركة/العملة في Inventory
                        item = Inventory.query.filter_by(item_name=currency, item_type='company', balance=None).first()
                        # نستخدم item_type='company' وitem_name=currency، ونخزن اسم الشركة في حقل آخر إذا كان متاحًا
                        # أو نستخدم منطقك الحالي للجرد
                        # هنا سنخزن باسم الشركة+العملة
                        item = Inventory.query.filter_by(item_name=f'{company_name}__{currency}', item_type='company').first()
                        if not item:
                            item = Inventory(item_name=f'{company_name}__{currency}', item_type='company', balance=0)
                            db.session.add(item)
                        item.balance = float(row[idx])
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/company/transaction', methods=['POST'])
def company_transaction():
    try:
        data = request.json
        company_name = data.get('company_name')
        operation = data.get('operation')  # 'withdraw' or 'deposit'
        amount = float(data.get('amount', 0))
        currency = data.get('currency')
        if not company_name or not operation or not amount or not currency:
            return jsonify({'error': 'بيانات ناقصة'}), 400
        # رصيد الشركة في جدول Inventory
        company_item = Inventory.query.filter_by(item_name=f'{company_name}__{currency}', item_type='company').first()
        if not company_item:
            company_item = Inventory(item_name=f'{company_name}__{currency}', item_type='company', balance=0)
            db.session.add(company_item)
        # رصيد الجرد
        inventory_item = Inventory.query.filter_by(item_name=currency, item_type='currency').first()
        if not inventory_item:
            inventory_item = Inventory(item_name=currency, balance=0, item_type='currency')
            db.session.add(inventory_item)
        if operation == 'withdraw':
            company_item.balance -= amount
            inventory_item.balance += amount
        elif operation == 'deposit':
            company_item.balance += amount
            inventory_item.balance -= amount
        else:
            return jsonify({'error': 'نوع العملية غير صحيح'}), 400
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@inventory_bp.route('/amanat', methods=['GET'])
def get_amanat():
    amanat = Amanah.query.order_by(Amanah.date.desc()).all()
    result = []
    for a in amanat:
        result.append({
            'id': a.id,
            'type': a.type,
            'person_name': a.person_name,
            'currency': a.currency,
            'amount': a.amount,
            'date': a.date,
            'notes': a.notes
        })
    return jsonify({'amanat': result})

@inventory_bp.route('/amanat/add', methods=['POST'])
def add_amanah():
    data = request.json
    a = Amanah(
        type=data.get('type'),
        person_name=data.get('person_name'),
        currency=data.get('currency'),
        amount=data.get('amount'),
        date=data.get('date'),
        notes=data.get('notes')
    )
    db.session.add(a)
    db.session.commit()
    return jsonify({'success': True, 'id': a.id})

@inventory_bp.route('/account_statement/<currency>', methods=['GET'])
def account_statement(currency):
    entries = []
    # 1. حركات الصندوق (cashbox)
    for entry in CashBoxEntry.query.filter_by(currency=currency).all():
        entries.append({
            'date': entry.entry_date,
            'action': 'إيداع في الصندوق',
            'amount': entry.amount,
            'source': 'الصندوق',
            'notes': ''
        })
    # 2. الديون
    for debt in Debt.query.filter_by(currency=currency).all():
        if debt.type == 'borrowed_from_someone':
            entries.append({
                'date': debt.debt_date,
                'action': 'دين عليّ (سحب)',
                'amount': -debt.amount,
                'source': 'الديون',
                'notes': debt.notes
            })
        else:
            entries.append({
                'date': debt.debt_date,
                'action': 'دين لي (إيداع)',
                'amount': debt.amount,
                'source': 'الديون',
                'notes': debt.notes
            })
    # 3. الذهب (بيع/شراء)
    for g in GoldTransaction.query.filter_by(currency=currency).all():
        if g.type == 'buy':
            entries.append({
                'date': g.transaction_date,
                'action': 'شراء ذهب (سحب)',
                'amount': -g.amount,
                'source': 'الذهب',
                'notes': ''
            })
        else:
            entries.append({
                'date': g.transaction_date,
                'action': 'بيع ذهب (إيداع)',
                'amount': g.amount,
                'source': 'الذهب',
                'notes': ''
            })
    # 4. الحوالات
    for r in Remittance.query.filter_by(currency=currency).all():
        if r.type == 'send':
            entries.append({
                'date': r.remittance_date,
                'action': 'حوالة صادرة (سحب)',
                'amount': -r.amount,
                'source': 'الحوالات',
                'notes': r.notes
            })
        else:
            entries.append({
                'date': r.remittance_date,
                'action': 'حوالة واردة (إيداع)',
                'amount': r.amount,
                'source': 'الحوالات',
                'notes': r.notes
            })
    # 5. ذهب المحل
    for s in ShopGold.query.filter_by(piece_type=currency).all():
        entries.append({
            'date': s.entry_date,
            'action': 'ذهب محل (غير نقدي)',
            'amount': 0,
            'source': 'ذهب المحل',
            'notes': s.notes
        })
    # ترتيب حسب التاريخ
    entries.sort(key=lambda x: str(x['date']))
    total_in = sum(e['amount'] for e in entries if e['amount'] > 0)
    total_out = -sum(e['amount'] for e in entries if e['amount'] < 0)
    balance = total_in - total_out
    # تحويل التاريخ لنص
    for e in entries:
        if isinstance(e['date'], datetime):
            e['date'] = e['date'].strftime('%Y-%m-%d')
        else:
            e['date'] = str(e['date'])
    return jsonify({
        'entries': entries,
        'total_in': total_in,
        'total_out': total_out,
        'balance': balance
    })

